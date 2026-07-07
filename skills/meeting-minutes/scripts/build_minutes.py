#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""회의록 JSON 데이터를 템플릿(assets/회의록_템플릿.xlsx)에 채워 엑셀 회의록을 생성한다.

사용법:
    python build_minutes.py <data.json> [-o 출력.xlsx] [-t 템플릿.xlsx]

data.json 스키마 (모든 문자열, 날짜는 "YYYY-MM-DD" 문자열, 모르면 "미정(확인 필요)"):
{
  "project": "프로젝트명",
  "meeting_name": "회의명",
  "datetime": "2026-07-06 14:00 ~ 15:30",
  "writer": "작성자",
  "attendees": ["[고객사명] 홍길동, 김철수", "[수행사명] 이영희, 박민수"],
  "reference_docs": ["문서명1", "문서명2"],
  "discussions": [
    {"category": "구분", "items": "1. 항목\n2. 항목", "content": "상세 내용", "note": "비고"}
  ],
  "todos": [
    {"category": "항목", "side": "영역(누가)", "content": "내용",
     "request_date": "", "due_date": "", "done_date": "", "assignee": ""}
  ],
  "issues": [
    {"category": "항목", "side": "영역", "content": "내용 및 처리방안",
     "status": "현황", "occur_date": "", "plan_date": "", "done_date": ""}
  ],
  "sheet_name": "2026.07.06"   // 생략 시 datetime의 날짜로 자동
}
"""
import argparse
import json
import math
import re
import sys
from copy import copy
from pathlib import Path

import openpyxl

# 템플릿 고정 좌표
ROW_DISC_PROTO = 14
ROW_TODO_TITLE = 16
ROW_TODO_PROTO = 18
ROW_ISSUE_TITLE = 20
ROW_ISSUE_PROTO = 22
COLS = "BCDEFGH"

# 열별 대략적 글자 수(한글 기준) — 행 높이 추정용
CHARS_PER_LINE = {"B": 6, "C": 9, "D": 33, "E": 6, "F": 6, "G": 6, "H": 6}
CHARS_MERGED_NOTE = 20  # 회의내용 비고(E:H 병합)


def err(msg):
    print(msg, file=sys.stderr)
    sys.exit(1)


def insert_rows_keep_merges(ws, at, n):
    """at 행 위치에 n행 삽입. at 이후에서 시작하는 병합 범위를 함께 아래로 민다."""
    if n <= 0:
        return
    shifted = []
    for m in list(ws.merged_cells.ranges):
        c1, r1, c2, r2 = m.bounds
        if r1 >= at:
            shifted.append((c1, r1, c2, r2))
            ws.unmerge_cells(str(m))
    ws.insert_rows(at, n)
    for c1, r1, c2, r2 in shifted:
        ws.merge_cells(start_row=r1 + n, start_column=c1,
                       end_row=r2 + n, end_column=c2)


def copy_row_style(ws, src_row, dst_row):
    for col in COLS:
        src = ws[f"{col}{src_row}"]
        dst = ws[f"{col}{dst_row}"]
        dst._style = copy(src._style)


def line_count(text, chars):
    if not text:
        return 1
    total = 0
    for line in str(text).split("\n"):
        total += max(1, math.ceil(len(line) / chars))
    return total


def set_data_row_height(ws, row, cells):
    """cells: {열문자: (텍스트, 줄당글자수)} — 가장 긴 셀 기준으로 행 높이 설정."""
    lines = max(line_count(t, c) for t, c in cells.values())
    ws.row_dimensions[row].height = max(19.5, lines * 13.8 + 4)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("data_json")
    ap.add_argument("-o", "--output", default=None)
    ap.add_argument("-t", "--template", default=None)
    args = ap.parse_args()

    data = json.loads(Path(args.data_json).read_text(encoding="utf-8-sig"))

    for key in ("project", "meeting_name", "datetime", "writer"):
        if not data.get(key):
            err(f"필수 항목 누락: {key}")

    script_dir = Path(__file__).resolve().parent
    template = Path(args.template) if args.template else script_dir.parent / "assets" / "회의록_템플릿.xlsx"
    if not template.exists():
        err(f"템플릿을 찾을 수 없음: {template}")

    wb = openpyxl.load_workbook(template)
    ws = wb.active

    # ── 회의 개요 ──
    ws["B2"] = f"[{data['project']}] 회의록"
    ws["C3"] = data["meeting_name"]
    ws["C4"] = data["datetime"]
    ws["C5"] = data["writer"]

    attendees = data.get("attendees") or []
    ws["C6"] = attendees[0] if attendees else None
    if len(attendees) > 1:
        ws["C7"] = "\n".join(attendees[1:])
        if len(attendees) > 2:
            ws.row_dimensions[7].height = 19.5 * (len(attendees) - 1)

    docs = data.get("reference_docs") or []
    for i, coord in enumerate(["C8", "C9", "C10", "C11"]):
        if i < 3 or len(docs) <= 4:
            ws[coord] = docs[i] if i < len(docs) else None
        else:
            ws[coord] = "\n".join(docs[3:])  # 5개 이상이면 마지막 칸에 합침
            ws.row_dimensions[11].height = 19.5 * len(docs[3:])

    discussions = data.get("discussions") or []
    todos = data.get("todos") or []
    issues = data.get("issues") or []

    # ── 행 삽입 (위에서 아래로, 오프셋 누적) ──
    nd, nt, ni = len(discussions), len(todos), len(issues)
    insert_rows_keep_merges(ws, ROW_DISC_PROTO + 1, nd - 1)
    off1 = max(0, nd - 1)
    insert_rows_keep_merges(ws, ROW_TODO_PROTO + off1 + 1, nt - 1)
    off2 = off1 + max(0, nt - 1)
    insert_rows_keep_merges(ws, ROW_ISSUE_PROTO + off2 + 1, ni - 1)

    # ── 회의 내용 ──
    for i, d in enumerate(discussions):
        r = ROW_DISC_PROTO + i
        if i > 0:
            copy_row_style(ws, ROW_DISC_PROTO, r)
            ws.merge_cells(f"E{r}:H{r}")
        ws[f"B{r}"] = d.get("category", "")
        ws[f"C{r}"] = d.get("items", "")
        ws[f"D{r}"] = d.get("content", "")
        ws[f"E{r}"] = d.get("note", "")
        set_data_row_height(ws, r, {
            "C": (d.get("items", ""), CHARS_PER_LINE["C"]),
            "D": (d.get("content", ""), CHARS_PER_LINE["D"]),
            "E": (d.get("note", ""), CHARS_MERGED_NOTE),
        })

    # ── TO-DO TASK ──
    base_todo = ROW_TODO_PROTO + off1
    for i, t in enumerate(todos):
        r = base_todo + i
        if i > 0:
            copy_row_style(ws, base_todo, r)
        ws[f"B{r}"] = t.get("category", "")
        ws[f"C{r}"] = t.get("side", "")
        ws[f"D{r}"] = t.get("content", "")
        ws[f"E{r}"] = t.get("request_date", "")
        ws[f"F{r}"] = t.get("due_date", "")
        ws[f"G{r}"] = t.get("done_date", "")
        ws[f"H{r}"] = t.get("assignee", "")
        set_data_row_height(ws, r, {
            "C": (t.get("side", ""), CHARS_PER_LINE["C"]),
            "D": (t.get("content", ""), CHARS_PER_LINE["D"]),
        })

    # ── ISSUE REPORT ──
    base_issue = ROW_ISSUE_PROTO + off2
    for i, s in enumerate(issues):
        r = base_issue + i
        if i > 0:
            copy_row_style(ws, base_issue, r)
        ws[f"B{r}"] = s.get("category", "")
        ws[f"C{r}"] = s.get("side", "")
        ws[f"D{r}"] = s.get("content", "")
        ws[f"E{r}"] = s.get("status", "")
        ws[f"F{r}"] = s.get("occur_date", "")
        ws[f"G{r}"] = s.get("plan_date", "")
        ws[f"H{r}"] = s.get("done_date", "")
        set_data_row_height(ws, r, {
            "C": (s.get("side", ""), CHARS_PER_LINE["C"]),
            "D": (s.get("content", ""), CHARS_PER_LINE["D"]),
        })

    # ── 시트명 ──
    sheet = data.get("sheet_name")
    if not sheet:
        m = re.search(r"(\d{4})-(\d{2})-(\d{2})", data["datetime"])
        sheet = f"{m.group(1)}.{m.group(2)}.{m.group(3)}" if m else "회의록"
    ws.title = sheet[:31]

    out = Path(args.output) if args.output else Path(f"[{data['project']}] 회의록_{sheet.replace('.', '')}.xlsx")
    wb.save(out)
    print(json.dumps({"output": str(out), "discussions": nd, "todos": nt, "issues": ni},
                     ensure_ascii=False))


if __name__ == "__main__":
    main()
