#!/usr/bin/env python3
# 요구사항 정의서(.xlsx) 생성 스크립트
# 〇〇 통합관리시트 '04_요구사항 정의서' 양식을 그대로 재현한다.
#
# 사용법:
#   python build_requirements_xlsx.py <input.json> <output.xlsx> [프로젝트명]
#
# 입력 JSON: 요구사항 dict의 리스트. 스키마는 references/template-spec.md 참조.
# 출력: 헤더 2줄 병합 + 제목행 + 틀고정(A5)까지 재현한 엑셀.
#
# 설계 원칙:
#  - 요구사항ID는 대분류별 접두어(CHT/COM/BOS/SYS/MNG…)로 독립 채번한다.
#  - 출처가 'AI추정'으로 시작하는 행은 검토가 필요하므로 배경색으로 표시한다.
#  - 결과(생성 건수·AI추정 건수·경고)는 JSON으로 stdout에, 에러는 stderr로 낸다.

import sys
import json
import re

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.stderr.write("openpyxl이 필요합니다: pip install openpyxl\n")
    sys.exit(1)

# ── 컬럼 정의 (실제 양식 순서 그대로) ───────────────────────────────
# (열번호, row3 헤더, row4 서브헤더, JSON 키, 너비)
COLUMNS = [
    (1,  "No",          None,       None,               4.7),
    (2,  "요구사항 ID",  None,       "요구사항ID",        11.6),
    (3,  "RFP 매핑 ID",  None,       "rfp_id",           9.2),
    (4,  "Q_N",         None,       "q_n",              9.0),
    (5,  "대분류",       None,       "대분류",            9.0),
    (6,  "중분류",       None,       "중분류",            15.1),
    (7,  "요구사항명",   "1Depth",   "요구사항명_1depth", 13.1),
    (8,  None,          "2Depth",   "요구사항명_2depth", 20.9),
    (9,  "요구사항 상세설명", None,   "상세설명",          49.6),
    (10, "요청자",       "소속",     "요청자_소속",       16.3),
    (11, None,          "파트",     "요청자_파트",       10.0),
    (12, None,          "이름",     "요청자_이름",       10.0),
    (13, None,          "요청일시", "요청일시",          14.0),
    (14, "유형",         None,       "유형",             15.0),
    (15, "출처",         None,       "출처",             22.8),
    (16, "우선순위",     None,       "우선순위",          8.0),
    (17, "난이도",       None,       "난이도",           8.0),
    (18, "수용 여부",    None,       "수용여부",          12.4),
    (19, "담당자",       None,       "담당자",           7.8),
    (20, "처리방안",     None,       "처리방안",          36.1),
    (21, "관련 자료",    None,       "관련자료",          20.8),
]

# 병합 헤더 (요구사항명 G3:H3, 요청자 J3:M3), 나머지는 3~4행 세로 병합
MERGE_HGROUP = {7: 8, 10: 13}  # 시작열 -> 끝열 (가로 병합)

# 대분류 → 요구사항ID 접두어
PREFIX_MAP = {
    "챗봇": "CHT",
    "공통": "COM",
    "어드민": "BOS",
    "시스템": "SYS",
    "관리": "MNG",
}

HEADER_FILL = PatternFill("solid", fgColor="FF000000")
HEADER_FONT = Font(bold=True, color="FFFFFFFF", size=10, name="맑은 고딕")
TITLE_FONT = Font(bold=True, size=13, name="맑은 고딕")
BODY_FONT = Font(size=10, name="맑은 고딕")
AI_FILL = PatternFill("solid", fgColor="FFFFF2CC")  # AI추정 행 강조(연노랑)
THIN = Side(style="thin", color="FFBFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def derive_prefix(대분류):
    if not 대분류:
        return "ETC"
    return PREFIX_MAP.get(str(대분류).strip(), "ETC")


def assign_ids(reqs):
    """대분류별로 REQ-<PREFIX>-### 채번. 이미 ID가 있으면 존중하되 중복은 경고."""
    counters = {}
    seen = set()
    warnings = []
    for r in reqs:
        existing = str(r.get("요구사항ID") or "").strip()
        prefix = derive_prefix(r.get("대분류"))
        if existing:
            if existing in seen:
                warnings.append(f"요구사항ID 중복: {existing}")
            seen.add(existing)
            # 기존 ID의 번호를 카운터에 반영해 이어치기
            m = re.match(r"REQ-([A-Z]+)-(\d+)", existing)
            if m:
                p, n = m.group(1), int(m.group(2))
                counters[p] = max(counters.get(p, 0), n)
            continue
        counters[prefix] = counters.get(prefix, 0) + 1
        new_id = f"REQ-{prefix}-{counters[prefix]:03d}"
        while new_id in seen:  # 안전장치
            counters[prefix] += 1
            new_id = f"REQ-{prefix}-{counters[prefix]:03d}"
        r["요구사항ID"] = new_id
        seen.add(new_id)
    return warnings


def build(reqs, out_path, project_name=""):
    warnings = assign_ids(reqs)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "04_요구사항 정의서"

    # 제목행 (row1)
    title = "요구사항 정의서"
    if project_name:
        title = f"요구사항 정의서 — {project_name}"
    ws.cell(1, 1, title).font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLUMNS))
    ws.cell(1, 1).alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 33.75

    # 헤더 (row3~4)
    for colnum, h3, h4, _key, width in COLUMNS:
        if h3 is not None:
            c = ws.cell(3, colnum, h3)
            c.fill = HEADER_FILL
            c.font = HEADER_FONT
            c.alignment = CENTER
            c.border = BORDER
        if h4 is not None:
            c = ws.cell(4, colnum, h4)
            c.fill = HEADER_FILL
            c.font = HEADER_FONT
            c.alignment = CENTER
            c.border = BORDER
        ws.column_dimensions[get_column_letter(colnum)].width = width

    # 헤더 병합: 가로 그룹(요구사항명/요청자) + 나머지 세로(3:4)
    hgroup_cols = set()
    for start, end in MERGE_HGROUP.items():
        ws.merge_cells(start_row=3, start_column=start, end_row=3, end_column=end)
        for c in range(start, end + 1):
            hgroup_cols.add(c)
    for colnum, h3, h4, _key, _w in COLUMNS:
        if colnum in hgroup_cols:
            continue
        ws.merge_cells(start_row=3, start_column=colnum, end_row=4, end_column=colnum)
    # 헤더 병합칸 테두리/채움 보정 (row4의 병합 셀도 검정 유지)
    for colnum, h3, h4, _key, _w in COLUMNS:
        for row in (3, 4):
            cell = ws.cell(row, colnum)
            if cell.fill.patternType is None:
                cell.fill = HEADER_FILL
            cell.border = BORDER
    ws.row_dimensions[3].height = 30

    # 데이터 (row5~)
    ai_count = 0
    row = 5
    for idx, r in enumerate(reqs, start=1):
        출처 = str(r.get("출처") or "")
        is_ai = 출처.replace(" ", "").startswith("AI추정")
        if is_ai:
            ai_count += 1
        ws.cell(row, 1, idx)  # No
        for colnum, _h3, _h4, key, _w in COLUMNS:
            if key is None or colnum == 1:
                continue
            val = r.get(key, "")
            cell = ws.cell(row, colnum, val if val is not None else "")
            cell.font = BODY_FONT
            cell.border = BORDER
            cell.alignment = LEFT if colnum == 9 else CENTER
            if is_ai and colnum == 15:  # 출처 셀 강조
                cell.fill = AI_FILL
        ws.cell(row, 1).font = BODY_FONT
        ws.cell(row, 1).border = BORDER
        ws.cell(row, 1).alignment = CENTER
        row += 1

    ws.freeze_panes = "A5"

    wb.save(out_path)

    ai_ratio = round(ai_count / len(reqs) * 100, 1) if reqs else 0
    result = {
        "output": out_path,
        "count": len(reqs),
        "ai_estimated": ai_count,
        "ai_ratio_percent": ai_ratio,
        "warnings": warnings,
    }
    if ai_ratio >= 30:
        result["warnings"].append(
            f"AI추정 비율이 {ai_ratio}%입니다. 입력 자료가 부족할 수 있으니 검토·보강을 권장합니다."
        )
    return result


def main():
    if len(sys.argv) < 3:
        sys.stderr.write("사용법: python build_requirements_xlsx.py <input.json> <output.xlsx> [프로젝트명]\n")
        sys.exit(1)
    in_path, out_path = sys.argv[1], sys.argv[2]
    project = sys.argv[3] if len(sys.argv) > 3 else ""
    try:
        with open(in_path, "r", encoding="utf-8") as f:
            data = json.load(f)
    except Exception as e:
        sys.stderr.write(f"입력 JSON 읽기 실패: {e}\n")
        sys.exit(1)
    reqs = data.get("requirements") if isinstance(data, dict) else data
    if not isinstance(reqs, list) or not reqs:
        sys.stderr.write("요구사항 목록(requirements)이 비어 있습니다.\n")
        sys.exit(1)
    result = build(reqs, out_path, project)
    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
