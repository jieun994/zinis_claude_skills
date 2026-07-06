# -*- coding: utf-8 -*-
"""
IA(사이트 구조 및 기능정의) 엑셀 생성 스크립트.

입력 JSON으로 받은 메뉴 트리를 엑셀 정보구조도/기능정의서로 만든다.
스타일은 wbs-generator 샘플의 디자인 언어를 따른다:
  - 컬럼 헤더 = 검정 배경 + 흰 글씨 / 그룹 헤더 = 진회색
  - 행 색: Depth1=진회색(흰글씨) · Depth2=연노랑 · Depth3+=흰색
  - 폰트 Malgun Gothic, 얇은 회색 테두리
시트 구성: 표지 → 사용안내 → 개정이력 → (Front/Admin 등) IA 시트들.

컬럼은 사용자가 고른 것만 표시(가변). Depth는 최대 5단계까지 유연하게.
Front/Admin을 시트로 나눌 수 있고, Admin 시트엔 역할별 권한(O/X) 칸을 둔다.

사용법:
  python build_ia.py <입력.json> <출력.xlsx>

입력 JSON 구조는 README.md / SKILL.md 참조.
"""
import sys
import json
import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FONT = "Malgun Gothic"

# ---- 색상 팔레트 (모던 슬레이트/네이비 + 블루 포인트) ----
C_PRIMARY = "FF1E293B"    # 타이틀 밴드 / 컬럼 헤더 (slate-800)
C_GROUP = "FF475569"      # 그룹 헤더 (slate-600)
C_ACCENT = "FF2563EB"     # 포인트 — 체크/● (blue-600)
C_ZEBRA = "FFF6F8FB"      # 옅은 줄무늬
C_TOTAL = "FFE8EEF7"      # 총계 행 (옅은 블루그레이)
C_PHASE2 = "FFEDEFF2"     # 2차 행 (옅은 회색)
C_XCELL = "FFF1F5F9"      # 권한 미허용 셀 (slate-100)
C_SECTION = "FFEEF2F8"    # 사용안내 섹션 헤더
C_SEC_TXT = "FF1E293B"
C_LABEL = "FF64748B"      # 라벨 글씨 (slate-500)
C_VALUE = "FF0F172A"      # 본문 진한 글씨 (slate-900)
C_MUTE = "FF94A3B8"       # 흐린 글씨 (slate-400)
WHITE = "FFFFFFFF"

_thin = Side(style="thin", color="FFE2E8F0")        # 기본 테두리 (slate-200)
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
_div = Side(style="medium", color="FFCBD5E1")       # 섹션 구분선 (slate-300)
_bot = Side(style="thin", color="FFCBD5E1")

OPT_LABELS = {
    "code": "Code",
    "tab": "Tab/서브 플로우",
    "page_def": "페이지 정의 / 요구사항",
    "detail": "세부 사항",
    "page": "Page",
    "note": "비고",
}
DEV_KEYS = ["디자인", "퍼블리싱", "개발"]
OPS = ["C", "R", "U", "D"]          # 권한 CRUD
GRP, HDR, FIRST = 3, 4, 5   # 그룹헤더 / 컬럼헤더 / 데이터 시작 행


def fill(c):
    return PatternFill("solid", fgColor=c)


def to_date(s):
    if not s:
        return None
    if isinstance(s, (datetime.datetime, datetime.date)):
        return s
    return datetime.datetime.strptime(str(s)[:10], "%Y-%m-%d")


# =====================================================================
#  IA 시트
# =====================================================================
def build_layout(cols, max_depth, roles):
    """시트 컬럼 레이아웃 [(key, label), ...]. Code(첫 열)/Depth/비고는 항상 포함."""
    layout = [("code", OPT_LABELS["code"])]
    for d in range(1, max_depth + 1):
        layout.append((f"depth{d}", f"Depth{d}"))
    if "tab" in cols:
        layout.append(("tab", OPT_LABELS["tab"]))
    if "page_def" in cols:
        layout.append(("page_def", OPT_LABELS["page_def"]))
    if "detail" in cols:
        layout.append(("detail", OPT_LABELS["detail"]))
    if "page" in cols:
        layout.append(("page", OPT_LABELS["page"]))
    if "dev_split" in cols:
        for dk in DEV_KEYS:
            layout.append((f"dev_{dk}", dk))
    for rn in (roles or []):                 # 권한: 역할마다 C/R/U/D 4칸
        for op in OPS:
            layout.append((f"auth_{rn}_{op}", op))
    layout.append(("note", OPT_LABELS["note"]))
    return layout


WIDTHS = {"code": 13, "tab": 18, "page_def": 30, "detail": 26,
          "page": 6, "note": 17}
DEPTH_W = {1: 13.5, 2: 15, 3: 30, 4: 22, 5: 18}


def col_width(key):
    if key.startswith("depth"):
        return DEPTH_W.get(int(key[5:]), 16)
    if key.startswith("dev_"):
        return 8
    if key.startswith("auth_"):
        return 4.5
    return WIDTHS.get(key, 14)


def crud_set(val):
    """auth 값을 허용 CRUD 집합으로 정규화한다."""
    if val is None:
        return set()
    if isinstance(val, dict):
        return {op for op in OPS
                if str(val.get(op, "")).upper() in ("O", "Y", "TRUE", "1", "●")}
    if isinstance(val, (list, tuple)):
        return {str(x).upper()[:1] for x in val} & set(OPS)
    s = str(val).strip().upper()
    if s in ("O", "ALL", "CRUD"):
        return set(OPS)
    if s in ("", "X", "-", "NONE"):
        return set()
    return {ch for ch in s if ch in OPS}


def _group_header(ws, label, c1, c2):
    if c2 > c1:
        ws.merge_cells(start_row=GRP, start_column=c1, end_row=GRP, end_column=c2)
    g = ws.cell(row=GRP, column=c1, value=label)
    g.font = Font(name=FONT, size=9, bold=True, color=WHITE)
    g.alignment = Alignment(horizontal="center", vertical="center")
    thin_w = Side(style="thin", color="FF64748B")
    for c in range(c1, c2 + 1):
        cell = ws.cell(row=GRP, column=c)
        cell.fill = fill(C_GROUP)
        cell.border = Border(left=thin_w, right=thin_w, top=thin_w, bottom=thin_w)


def render_ia_sheet(ws, sheet, project, default_cols, max_depth):
    ws.sheet_view.showGridLines = False        # 눈금선 끄기(기본)
    cols = sheet.get("columns", default_cols)
    roles = sheet.get("roles", [])
    layout = build_layout(cols, max_depth, roles)
    ncol = len(layout)
    kc = {k: i + 1 for i, (k, _) in enumerate(layout)}

    for i, (k, _) in enumerate(layout):
        ws.column_dimensions[get_column_letter(i + 1)].width = col_width(k)

    # 타이틀 밴드 (1행, 표 너비 전체 병합) — 네이비 배경 + 흰 글씨
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol)
    name = project.get("name", "")
    doc = sheet.get("title", "정보구조도")
    band = ws.cell(row=1, column=1)
    band.value = f"{name}   |   {doc}" if name else doc
    band.font = Font(name=FONT, size=14, bold=True, color=WHITE)
    band.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    for c in range(1, ncol + 1):
        ws.cell(row=1, column=c).fill = fill(C_PRIMARY)
    ws.row_dimensions[1].height = 34
    ws.row_dimensions[2].height = 6     # 밴드와 표 사이 여백

    # 그룹 헤더 (3행) — 개발구분 / 역할별 권한
    dev_cols = [kc[f"dev_{dk}"] for dk in DEV_KEYS if f"dev_{dk}" in kc]
    if dev_cols:
        _group_header(ws, "개발구분", min(dev_cols), max(dev_cols))
    for rn in roles:
        rc = [kc[f"auth_{rn}_{op}"] for op in OPS]
        _group_header(ws, rn, min(rc), max(rc))

    # 컬럼 헤더 (4행) — 네이비 배경 + 흰 글씨
    for i, (k, label) in enumerate(layout):
        c = ws.cell(row=HDR, column=i + 1, value=label)
        c.font = Font(name=FONT, size=9.5, bold=True, color=WHITE)
        c.alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
        c.fill = fill(C_PRIMARY)
        c.border = BORDER
    ws.row_dimensions[HDR].height = 30

    # 데이터 행
    rows = sheet.get("rows", [])
    r = FIRST
    di = 0
    for item in rows:
        depth = int(item.get("depth", 1))
        phase2 = bool(item.get("phase2"))
        is_d1 = depth == 1
        if phase2:
            row_fill = C_PHASE2
        elif is_d1:
            row_fill = WHITE            # 대분류는 흰색(아래 구분선으로 강조)
        else:
            row_fill = C_ZEBRA if di % 2 else WHITE
        txt = C_MUTE if phase2 else C_VALUE

        dkey = f"depth{min(depth, max_depth)}"
        if dkey in kc:
            ws.cell(row=r, column=kc[dkey], value=item.get("name", ""))
        for k in ("code", "tab", "page_def", "detail", "note"):
            if k in kc and item.get(k) not in (None, ""):
                ws.cell(row=r, column=kc[k], value=item.get(k))
        if "page" in kc and item.get("page") not in (None, ""):
            ws.cell(row=r, column=kc["page"], value=item.get("page"))
        dev = item.get("dev") or []
        auth = item.get("auth") or {}
        allowed_by_role = {rn: crud_set(auth.get(rn)) for rn in roles}

        # 서식 적용
        for cc in range(1, ncol + 1):
            cell = ws.cell(row=r, column=cc)
            cell.border = BORDER
            cell.fill = fill(row_fill)
            key = layout[cc - 1][0]
            cell.font = Font(name=FONT, size=9, bold=is_d1, color=txt)
            # 정렬
            if key.startswith("dev_") or key.startswith("auth_") \
                    or key in ("page", "code"):
                cell.alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrapText=True)
            # 개발구분 ● (포인트 색)
            if key.startswith("dev_"):
                dk = key[4:]
                if dk in dev:
                    cell.value = "●"
                    cell.font = Font(name=FONT, size=9, color=C_ACCENT)
            # 권한 CRUD — 허용은 ✓(포인트색), 미허용은 연회색 셀
            if key.startswith("auth_"):
                rn, op = key[5:].rsplit("_", 1)
                if op in allowed_by_role.get(rn, set()):
                    cell.value = "✓"
                    cell.font = Font(name=FONT, size=9, bold=True, color=C_ACCENT)
                elif not phase2:
                    cell.fill = fill(C_XCELL)
            # 대분류 행 위에 섹션 구분선
            if is_d1 and r > FIRST:
                cell.border = Border(left=_thin, right=_thin, top=_div, bottom=_thin)
        ws.row_dimensions[r].height = 19
        di += 1
        r += 1

    last_data = r - 1

    # 총계 행 (Page 컬럼이 있을 때)
    if "page" in kc and last_data >= FIRST:
        tr = r
        lab_col = kc.get("depth1", 1)
        lc = ws.cell(row=tr, column=lab_col, value="총 계")
        lc.font = Font(name=FONT, size=9.5, bold=True, color=C_VALUE)
        lc.alignment = Alignment(horizontal="center", vertical="center")
        pcol = kc["page"]
        pl = get_column_letter(pcol)
        sc = ws.cell(row=tr, column=pcol, value=f"=SUM({pl}{FIRST}:{pl}{last_data})")
        sc.font = Font(name=FONT, size=9.5, bold=True, color=C_ACCENT)
        sc.alignment = Alignment(horizontal="center", vertical="center")
        top_med = Side(style="medium", color="FF94A3B8")
        for cc in range(1, ncol + 1):
            cell = ws.cell(row=tr, column=cc)
            cell.border = Border(left=_thin, right=_thin, top=top_med, bottom=_thin)
            cell.fill = fill(C_TOTAL)
            if not cell.font or not cell.font.bold:
                cell.font = Font(name=FONT, size=9.5, bold=True, color=C_VALUE)
        ws.row_dimensions[tr].height = 20

    ws.freeze_panes = ws.cell(row=FIRST, column=1)


# =====================================================================
#  표지 / 사용안내 / 개정이력
# =====================================================================
def make_cover(ws, project, doc_type):
    ws.sheet_view.showGridLines = False
    widths = {"A": 2.5, "B": 14, "C": 14, "D": 14, "E": 14, "F": 14, "G": 14, "H": 2.5}
    for k, v in widths.items():
        ws.column_dimensions[k].width = v

    # 상단 액센트 밴드 (B2:G3)
    ws.merge_cells("B2:G3")
    for c in range(2, 8):
        for rr in (2, 3):
            ws.cell(row=rr, column=c).fill = fill(C_ACCENT)
    ws.row_dimensions[2].height = 10
    ws.row_dimensions[3].height = 10

    # 제목 블록 (B5:G9, 네이비 배경 + 흰 글씨)
    ws.merge_cells("B5:G9")
    for c in range(2, 8):
        for rr in range(5, 10):
            ws.cell(row=rr, column=c).fill = fill(C_PRIMARY)
    tc = ws.cell(row=5, column=2)
    tc.value = project.get("name", "")
    tc.font = Font(name=FONT, size=26, bold=True, color=WHITE)
    tc.alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
    for rr in range(5, 10):
        ws.row_dimensions[rr].height = 26

    # 문서 종류 부제 (B11:G11)
    ws.merge_cells("B11:G11")
    sc = ws.cell(row=11, column=2)
    sc.value = doc_type
    sc.font = Font(name=FONT, size=13, bold=True, color=C_LABEL)
    sc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[11].height = 22

    # 메타 정보 표 (라벨=연한 배경, 값=흰 배경, 카드 느낌)
    today = datetime.date.today().strftime("%Y-%m-%d")
    items = [
        ("작 성 자", project.get("author", "")),
        ("작 성 일", project.get("author_date") or today),
        ("버    전", project.get("version", "")),
    ]
    start = 16
    side = Side(style="thin", color="FFD7DEE8")
    bd = Border(left=side, right=side, top=side, bottom=side)
    for i, (lab, val) in enumerate(items):
        r = start + i
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=3)
        lc = ws.cell(row=r, column=3, value=lab)
        lc.font = Font(name=FONT, size=10, bold=True, color=C_LABEL)
        lc.fill = fill(C_SECTION)
        lc.alignment = Alignment(horizontal="center", vertical="center")
        lc.border = bd
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=6)
        vc = ws.cell(row=r, column=4, value=val)
        vc.font = Font(name=FONT, size=10, color=C_VALUE)
        vc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        for c in range(4, 7):
            ws.cell(row=r, column=c).border = bd
        ws.row_dimensions[r].height = 22


GUIDE = [
    ("이 파일은?", [
        "서비스의 정보구조(IA) — 메뉴 구조와 기능정의서입니다.",
        "시트 구성: 표지 · 사용안내 · 개정이력 · (Front/Admin 등) 정보구조도",
    ]),
    ("표 보는 법", [
        "Code = 화면 코드(첫 열)  /  Depth1~5 = 메뉴 깊이 (한 칸씩 들여쓰기)",
        "페이지 정의/요구사항 = 그 화면의 유형(디자인 페이지·게시판·폼 등)",
        "세부 사항 = 화면 안 항목·분기  /  Page = 화면 수 (맨 아래 총계 자동 합산)",
    ]),
    ("개발구분 ● 의 의미", [
        "각 화면을 누가 작업하는지 표시: 디자인 / 퍼블리싱 / 개발",
        "● = 해당 직무가 작업하는 화면. 페이지 유형 보고 자동 제안된 값이니 자유롭게 고치세요.",
    ]),
    ("권한 (Admin 시트) — 역할별 CRUD", [
        "역할(예: 조사원·지방실사·관리자)마다 4개 칸으로 권한을 표시합니다.",
        "C=등록(Create) · R=조회(Read) · U=수정(Update) · D=삭제(Delete)",
        "O = 가능한 동작.  비어 있는(연회색) 칸 = 불가.  예) 조회만 = R 칸만 O.",
    ]),
    ("행 색의 의미", [
        "기본은 모두 흰 배경 · 검은 글씨입니다 (Depth와 무관).",
        "회색 행 = 2차(다음 단계에 만들 기능). 비고에 '2차'로 표시됩니다.",
    ]),
    ("직접 수정할 때", [
        "메뉴를 추가하려면 같은 레벨의 행을 복사·붙여넣기 하면 서식이 따라옵니다.",
        "Page 총계는 수식이라 행을 더하거나 지워도 자동으로 다시 계산됩니다.",
    ]),
]


def make_guide(ws, doc_type):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 112
    t = ws.cell(row=1, column=2, value=f"{doc_type} 사용 안내")
    t.font = Font(name=FONT, size=16, bold=True)
    t.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 26
    r = 3
    for head, lines in GUIDE:
        hc = ws.cell(row=r, column=2, value=head)
        hc.font = Font(name=FONT, size=11, bold=True, color=C_SEC_TXT)
        hc.fill = fill(C_SECTION)
        hc.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[r].height = 20
        r += 1
        for ln in lines:
            bc = ws.cell(row=r, column=2, value=ln)
            bc.font = Font(name=FONT, size=10)
            bc.alignment = Alignment(horizontal="left", vertical="center")
            r += 1
        r += 1


def make_revision(ws, project):
    ws.sheet_view.showGridLines = False
    for k, v in {"A": 10, "B": 14, "C": 12, "D": 55}.items():
        ws.column_dimensions[k].width = v
    t = ws.cell(row=1, column=1, value="개정이력")
    t.font = Font(name=FONT, size=16, bold=True, color=C_VALUE)
    # 헤더 (3행) — 네이비 배경 + 흰 글씨
    for i, h in enumerate(["버전", "날짜", "작성자", "변경 내용"]):
        c = ws.cell(row=3, column=i + 1, value=h)
        c.font = Font(name=FONT, size=10, bold=True, color=WHITE)
        c.fill = fill(C_PRIMARY)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER
    ws.row_dimensions[3].height = 24
    # 최초 작성 행 (4행)
    today = datetime.date.today().strftime("%Y-%m-%d")
    vals = [project.get("version", ""), project.get("author_date") or today,
            project.get("author", ""), "최초 작성"]
    for i, v in enumerate(vals):
        c = ws.cell(row=4, column=i + 1, value=v)
        c.font = Font(name=FONT, size=10, color=C_VALUE)
        c.alignment = Alignment(horizontal="left" if i == 3 else "center", vertical="center")
        c.border = BORDER
    ws.row_dimensions[4].height = 20


# =====================================================================
def build(input_json, output_xlsx):
    with open(input_json, encoding="utf-8") as f:
        data = json.load(f)
    project = data.get("project", {})
    default_cols = data.get("columns", [])
    max_depth = int(data.get("max_depth", 5))
    sheets = data.get("sheets", [])
    if not sheets:
        raise SystemExit("sheets 가 비어 있습니다. 최소 1개 시트가 필요합니다.")
    doc_type = data.get("doc_type") or sheets[0].get("title") or "정보구조도(IA)"

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    make_cover(wb.create_sheet("표지"), project, doc_type)
    make_guide(wb.create_sheet("사용안내"), doc_type)
    make_revision(wb.create_sheet("개정이력"), project)

    for sheet in sheets:
        name = sheet.get("sheet_name") or sheet.get("title") or "IA"
        ws = wb.create_sheet(title=name[:31])
        render_ia_sheet(ws, sheet, project, default_cols, max_depth)

    wb.save(output_xlsx)
    total = sum(len(s.get("rows", [])) for s in sheets)
    print(f"완료: {output_xlsx}  (IA 시트 {len(sheets)}개, 메뉴 {total}개 + 표지/사용안내/개정이력)")


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("사용법: python build_ia.py <입력.json> <출력.xlsx>")
        sys.exit(1)
    build(sys.argv[1], sys.argv[2])
