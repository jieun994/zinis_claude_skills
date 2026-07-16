# -*- coding: utf-8 -*-
"""qa-master 테스트케이스 엑셀 생성 공통 모듈

사용법:
    import sys
    sys.path.insert(0, r"<이 스킬의 scripts 폴더 경로>")
    from make_tc_excel import build_workbook

    build_workbook(overview, scenarios, testcases, out_path)
    build_workbook(overview, scenarios, testcases, out_path, submission=True)  # 제출용

인자:
    overview   : [(항목, 내용), ...]                      — 개요 시트 행
    scenarios  : [[시나리오ID, 구분, 시나리오명, 설명, 연관TC수, 근거], ...]
    testcases  : [[TC_HEADER 19개 순서의 값], ...]
                 실행 기록 컬럼(실제 결과~비고)은 빈 문자열, 실행결과는 "미실행" 권장
    out_path   : 저장 경로 (.xlsx)
    submission : True면 제출용 — 적용 기법 컬럼 삭제, 출처의 "AI보완(...)"을 "추가 도출"로 치환
"""
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

TC_HEADER = ["TC ID", "대분류", "중분류", "소분류", "시나리오 ID", "우선순위", "적용 기법",
             "사전조건", "테스트 절차", "입력 데이터", "기대결과", "출처",
             "실행결과", "실제 결과", "재현 절차", "발견 환경", "상태", "결함 ID", "비고"]
TC_WIDTHS = [13, 9, 12, 22, 13, 8, 16, 18, 34, 22, 34, 18, 10, 24, 24, 14, 11, 9, 20]
TC_CENTER_COLS = {"TC ID", "대분류", "시나리오 ID", "우선순위", "실행결과", "상태", "결함 ID"}
SCN_HEADER = ["시나리오 ID", "구분(대분류)", "시나리오명", "설명(사용자 관점 흐름)", "연관 TC 수", "근거(출처)"]
SCN_WIDTHS = [14, 12, 24, 55, 10, 24]
SCN_CENTER_COLS = {"시나리오 ID", "구분(대분류)", "연관 TC 수"}

HEADER_FILL = PatternFill("solid", fgColor="4A4A4A")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
BODY_FONT = Font(size=10)
HIGH_FILL = PatternFill("solid", fgColor="FCE4E4")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="top", wrap_text=True)


def _style_table(ws, header, widths, center_cols, freeze=True):
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER
    for r in range(2, ws.max_row + 1):
        for c in range(1, len(header) + 1):
            cell = ws.cell(r, c)
            cell.font = BODY_FONT
            cell.border = BORDER
            cell.alignment = CENTER if header[c - 1] in center_cols else WRAP
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    if freeze:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(header))}{ws.max_row}"


def build_workbook(overview, scenarios, testcases, out_path, submission=False):
    header = list(TC_HEADER)
    widths = list(TC_WIDTHS)
    tcs = [list(row) for row in testcases]

    if submission:
        idx_tech = header.index("적용 기법")
        idx_src = header.index("출처")
        header.pop(idx_tech)
        widths.pop(idx_tech)
        for row in tcs:
            row.pop(idx_tech)
        idx_src -= 1  # 적용 기법 삭제로 한 칸 당겨짐
        for row in tcs:
            row[idx_src] = re.sub(r"AI보완\([^)]*\)", "추가 도출", str(row[idx_src]))

    wb = Workbook()

    # 시트 1: 개요 (세로 2열)
    ws = wb.active
    ws.title = "개요"
    ws.append(["항목", "내용"])
    for row in overview:
        ws.append(list(row))
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER
    for r in range(2, ws.max_row + 1):
        ws.cell(r, 1).font = Font(size=10, bold=True)
        ws.cell(r, 2).font = BODY_FONT
        for c in (1, 2):
            ws.cell(r, c).alignment = WRAP
            ws.cell(r, c).border = BORDER
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 90

    # 시트 2: 테스트 시나리오
    ws2 = wb.create_sheet("테스트 시나리오")
    ws2.append(SCN_HEADER)
    for row in scenarios:
        ws2.append(list(row))
    _style_table(ws2, SCN_HEADER, SCN_WIDTHS, SCN_CENTER_COLS)

    # 시트 3: 테스트케이스
    ws3 = wb.create_sheet("테스트케이스")
    ws3.append(header)
    for row in tcs:
        ws3.append(row)
    _style_table(ws3, header, widths, TC_CENTER_COLS)

    idx_pri = header.index("우선순위") + 1
    for r in range(2, ws3.max_row + 1):
        if ws3.cell(r, idx_pri).value == "상":
            ws3.cell(r, idx_pri).fill = HIGH_FILL

    col_result = get_column_letter(header.index("실행결과") + 1)
    dv1 = DataValidation(type="list", formula1='"Pass,Fail,Block,미실행"', allow_blank=True)
    ws3.add_data_validation(dv1)
    dv1.add(f"{col_result}2:{col_result}{max(ws3.max_row, 2)}")

    col_status = get_column_letter(header.index("상태") + 1)
    dv2 = DataValidation(type="list", formula1='"신규,수정중,수정완료,재확인완료"', allow_blank=True)
    ws3.add_data_validation(dv2)
    dv2.add(f"{col_status}2:{col_status}{max(ws3.max_row, 2)}")

    wb.save(out_path)
    return out_path
