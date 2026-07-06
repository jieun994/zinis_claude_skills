# -*- coding: utf-8 -*-
"""
WBS 빈 템플릿 생성기 — 원본 WBS 샘플을 베이스로 구조만 수정한다.
원본 표(좌측)의 스타일/서식은 그대로 보존하고:
  1) No를 3칸(L1/L2/L3)으로 (원본의 빈 No칸 B, L4 No칸 E 삭제)
  2) L4 이름 컬럼(I) 삭제
  3) 행 그룹(아웃라인)·숨김 전부 해제
  4) 데이터 행을 60개(7~66)만 남김
  5) 달력/간트를 새 컬럼 위치로 재구성(조건부서식 재생성), 공휴일은 '공휴일' 시트 참조
  + 시트명 표지/WBS/공휴일, 순서 표지·WBS·공휴일, 기준일=TODAY()

실행: python make_template.py <원본_WBS샘플.xlsx> <출력_template.xlsx>
"""
import sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule
from openpyxl.formatting.formatting import ConditionalFormattingList
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

FONT = "Malgun Gothic"
NROWS = 60
FIRST = 7
LAST = FIRST + NROWS - 1            # 66
CAL_START = 22                      # 삭제 후 달력 시작(V)
NDAYS = 370

C_HEADER = "FF000000"
C_HOLIDAY = "FFE7E6E6"
C_DONE = "FF2563EB"
C_COMPLETE = "FF16A34A"      # 완료 = 초록
C_REMAIN = "FFBFDBFE"
C_WAIT = "FFCBD5E1"
C_HOLD = "FFD97706"
C_TODAY = "FFFCA5A5"

thin = Side(style="thin", color="FFD1D5DB")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

# 삭제 후 표 컬럼 너비 (1-base). A~E(No 3칸+L1,L2 이름)는 2 → 들여쓰기형
WIDTHS = {1: 2, 2: 2, 3: 2, 4: 2, 5: 2, 6: 30, 7: 16, 8: 10, 9: 10,
          10: 12, 11: 24, 12: 11, 13: 11, 14: 6, 15: 8, 16: 9, 17: 8,
          18: 9, 19: 8, 20: 9, 21: 16}
LEGEND = [("FF2563EB", "진행"), ("FF16A34A", "완료"), ("FFBFDBFE", "잔여"),
          ("FFCBD5E1", "대기"), ("FFD97706", "보류"), ("FFE7E6E6", "주말/공휴일"),
          ("FFFCA5A5", "기준일")]


def fill(c):
    # CF 조건부서식은 bgColor를 색으로 인식 → fg/bg 둘 다 채운다 (셀 칠에도 무해)
    return PatternFill("solid", fgColor=c, bgColor=c)


def build_guide_sheet(wb):
    """파일 안에 들어가는 '사용안내' 온보딩 시트."""
    if "사용안내" in wb.sheetnames:
        del wb["사용안내"]
    g = wb.create_sheet("사용안내")
    g.sheet_view.showGridLines = False
    g.column_dimensions["A"].width = 3
    g.column_dimensions["B"].width = 112
    LINES = [
        ("title", "WBS 사용 안내"),
        ("blank", ""),
        ("h", "이 파일은?"),
        ("b", "프로젝트 작업분해구조(WBS)입니다. 시트 구성: 표지 · 사용안내 · 개정이력 · 요약 · WBS · 공휴일"),
        ("blank", ""),
        ("h", "내가 직접 넣는 것 — 말단(L3) 작업"),
        ("b", "시작일 · 종료일 · 진척율(%) · 상태   ← 이 4가지만 입력하면 됩니다"),
        ("b", "→ 진행일 · 총/목표/실 작업량 · 상위 진척율 · 간트차트는 모두 자동 계산됩니다"),
        ("blank", ""),
        ("h", "상태와 간트 색 — '상태' 칸은 드롭다운"),
        ("b", "완료 = 초록   /   진행중·고객검토·수정반영 = 파랑+연파랑   /   대기 = 회청   /   보류 = 주황"),
        ("blank", ""),
        ("h", "직접 수정할 때"),
        ("b", "[안전] 시작일·종료일·진척율·상태 칸만 바꾸면 나머지(진행일·작업량·간트)는 자동으로 따라옵니다."),
        ("b", "[주의] 행을 추가하거나 지울 때는 수식 확인이 필요합니다:"),
        ("b2", "· 말단(L3) 추가 → 바로 위 같은 레벨 행을 복사·붙여넣기 (그 행엔 시작/종료/진척율/상태만 입력)"),
        ("b2", "· 상위(L1·L2)의 총작업량·실작업량은 '자식 칸을 더하는 수식'(예: =O9+O10)입니다."),
        ("b2", "  자식 작업을 추가하면 이 합계 수식에 새 행을 포함해 주세요 (예: =O9+O10+O11)."),
        ("blank", ""),
        ("h", "레벨별 계산 규칙 (참고 — 평소엔 몰라도 됩니다)"),
        ("b", "총작업량 :  L3 = 진행일        /  상위(L1·L2) = 하위 작업들의 합"),
        ("b", "실작업량 :  L3 = 총작업량 × 진척율  /  상위 = 하위 작업들의 합"),
        ("b", "진척율   :  L3 = 직접 입력      /  상위 = 실작업량 ÷ 총작업량"),
        ("b", "목표작업량 = 총작업량 × (기준일 가중치),  진행일 = 평일 수(공휴일 제외)  — 전부 자동"),
        ("blank", ""),
        ("h", "참고"),
        ("b", "· 총/목표/실 작업량 열은 계산용이라 기본 숨김입니다 (열 머리글에서 펼치면 보입니다)."),
        ("b", "· 기준일(오늘 선)과 공휴일(대한민국)은 자동으로 처리됩니다."),
    ]
    r = 1
    for kind, text in LINES:
        cell = g.cell(row=r, column=2, value=text)
        if kind == "title":
            cell.font = Font(name=FONT, size=16, bold=True)
            g.row_dimensions[r].height = 26
        elif kind == "h":
            cell.font = Font(name=FONT, size=11, bold=True, color="FF1F2937")
            cell.fill = PatternFill("solid", fgColor="FFF3F4F6")
            g.row_dimensions[r].height = 20
        elif kind == "b2":
            cell.font = Font(name=FONT, size=10, color="FF444444")
        else:
            cell.font = Font(name=FONT, size=10)
        cell.alignment = Alignment(horizontal="left", vertical="center")
        r += 1
    return g


def build_template(src, out):
    wb = openpyxl.load_workbook(src)
    ws = wb["01_WBS"]; ws.title = "WBS"
    cv = wb["문서표지"]; cv.title = "표지"

    for coord in ("B11", "D18", "D20", "D22", "D24"):
        cv[coord] = None
    cv["B37"] = None          # 하단 "〇〇 Co., Ltd." 제거
    cv._images = []           # 우측 상단 로고 이미지 제거

    # 데이터 값 비우기 (행7~157, A~X)
    for r in range(7, 158):
        for c in range(1, 25):
            if ws.cell(row=r, column=c).value is not None:
                ws.cell(row=r, column=c).value = None

    # (3) 그룹/숨김 해제
    for rd in ws.row_dimensions.values():
        rd.outline_level = 0
        rd.hidden = False
    for cd in ws.column_dimensions.values():
        cd.outline_level = 0
        cd.hidden = False

    # 기존 조건부서식/병합 전부 제거 (새로 깐다)
    ws.conditional_formatting = ConditionalFormattingList()
    for rng in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(rng))

    # (1,2) 열 삭제: L4이름(I=9) → L4No(E=5) → 빈No(B=2)  (오른쪽부터)
    ws.delete_cols(9)
    ws.delete_cols(5)
    ws.delete_cols(2)

    # (4) 행 60개만: 67行 이하 모두 삭제
    if ws.max_row > LAST:
        ws.delete_rows(LAST + 1, ws.max_row - LAST)

    # No 헤더 병합 A6:C6
    ws.merge_cells("A6:C6")

    # 메타: 기준일 값셀 = TODAY(), 시작/종료 값셀 비움(build가 채움)
    # 원본 P3(시작값)->M3, U3(종료값)->R3, W3(기준값)->T3  (열3개 삭제 반영)
    ws["M3"] = None      # 시작일 (build가 채움)
    ws["R3"] = None      # 종료일 (build가 채움)
    ws["T3"] = None      # 기준일 (build가 입력값으로 채움)
    ws["T3"].number_format = "yyyy-mm-dd"

    # 잔여 서식 정리
    white = PatternFill("solid", fgColor="FFFFFFFF")
    nofill = PatternFill(fill_type=None)
    # (#4) 데이터 행 7~66 표영역(A~U) 흰색으로 초기화 (원본 잔여 색 제거)
    for r in range(FIRST, LAST + 1):
        for c in range(1, 22):
            ws.cell(row=r, column=c).fill = white
    # (#5) 달력 영역(행2~66, 22열~) fill 제거 — 원본 검정 fill 잔여 삭제
    for r in range(2, LAST + 1):
        for c in range(CAL_START, CAL_START + NDAYS):
            cell = ws.cell(row=r, column=c)
            cell.fill = nofill
            if r <= 6:
                cell.value = None
    # (#7) 행3 달력영역 잔여 텍스트 제거 후 범례 삽입
    for c in range(CAL_START, CAL_START + NDAYS):
        ws.cell(row=3, column=c).value = None
    lc = CAL_START
    for color, label in LEGEND:
        box = ws.cell(row=3, column=lc)
        box.fill = PatternFill("solid", fgColor=color)
        box.border = BORDER
        ws.cell(row=3, column=lc + 1, value=label).font = Font(name=FONT, size=8)
        lc += 6
    # (#3) B4:C4 검은색
    for coord in ("B4", "C4"):
        ws[coord].fill = PatternFill("solid", fgColor="FF000000")

    # 상태 드롭다운 (T7:T66) — 대기/진행중/고객검토/수정반영/완료/보류
    ws.data_validations.dataValidation = []
    dv = DataValidation(type="list",
                        formula1='"대기,진행중,고객검토,수정반영,완료,보류"',
                        allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"T{FIRST}:T{LAST}")

    # 표 컬럼 너비
    for c, w in WIDTHS.items():
        ws.column_dimensions[get_column_letter(c)].width = w
    # 항상 숨김: 총/목표/실 작업량(15,16,17). 선택 칸(7,8,9)은 build가 제어.
    for c in (15, 16, 17):
        ws.column_dimensions[get_column_letter(c)].hidden = True
    # 잔여 하이퍼링크(예: flow.team) 전부 제거
    for row in ws.iter_rows():
        for cell in row:
            if cell.hyperlink is not None:
                cell.hyperlink = None
                if isinstance(cell.value, str) and "http" in cell.value:
                    cell.value = None
    try:
        ws._hyperlinks = []
    except Exception:
        pass

    # ---------- 달력: 수식 기반 (시작일 M3 / 종료일 R3 셀을 참조 → 시트에서 바꾸면 자동) ----------
    last_cal = CAL_START + NDAYS - 1
    sc = get_column_letter(CAL_START)
    lc = get_column_letter(last_cal)
    white = PatternFill("solid", fgColor="FFFFFFFF")
    for j in range(NDAYS):
        col = CAL_START + j
        L = get_column_letter(col)
        prev = get_column_letter(col - 1)
        ws.column_dimensions[L].width = 2.6
        # 행5 날짜: 첫칸=시작일($M$3), 이후=이전+1 (종료일 $R$3 넘으면 빈칸)
        c5 = ws.cell(row=5, column=col)
        c5.value = "=$M$3" if j == 0 else f'=IF(OR({prev}5="",{prev}5>=$R$3),"",{prev}5+1)'
        c5.number_format = "d"           # 날짜 = "일"만
        c5.font = Font(name=FONT, size=8)
        c5.alignment = Alignment(horizontal="center")
        c5.fill = white
        # 행6 요일: 날짜에서 자동
        c6 = ws.cell(row=6, column=col)
        c6.value = f'=IF({L}5="","",CHOOSE(WEEKDAY({L}5,2),"월","화","수","목","금","토","일"))'
        c6.font = Font(name=FONT, size=8, bold=True, color="FFFFFFFF")
        c6.fill = fill(C_HEADER)
        c6.alignment = Alignment(horizontal="center")
        # 행4 월: 매월 1일/첫칸에만 "n월" (build가 월별로 셀 병합 → 라벨 표시)
        c4 = ws.cell(row=4, column=col)
        c4.value = f'=IF({L}5="","",IF(OR({L}5=$M$3,DAY({L}5)=1),MONTH({L}5)&"월",""))'
        c4.font = Font(name=FONT, size=9, bold=True)
        c4.alignment = Alignment(horizontal="center")
        c4.fill = fill("FFD9D9D9")

    # ---------- 조건부서식(간트) 재생성 ----------
    # 표 컬럼(삭제 후): 시작일=L, 종료일=M, 진척율=S, 상태=T
    A = f"{sc}$5"
    I, J, P, Q = "$L7", "$M7", "$S7", "$T7"
    g = f'{I}<>"",{A}<>""'
    data_rng = f"{sc}{FIRST}:{lc}{LAST}"
    rules = [
        FormulaRule(formula=[f'AND({A}<>"",COUNTIF(공휴일!$A:$A,{A})>0)'], stopIfTrue=True, fill=fill(C_HOLIDAY)),
        FormulaRule(formula=[f'AND({A}<>"",WEEKDAY({A},2)>=6)'], stopIfTrue=True, fill=fill(C_HOLIDAY)),
        FormulaRule(formula=[f'AND({g},{A}>={I},{A}<={J},{Q}="완료")'], stopIfTrue=True, fill=fill(C_COMPLETE)),
        FormulaRule(formula=[f'AND({g},{A}>={I},{A}<={I}+({J}-{I})*{P},OR({Q}="진행중",{Q}="고객검토",{Q}="수정반영"))'], stopIfTrue=True, fill=fill(C_DONE)),
        FormulaRule(formula=[f'AND({g},{A}>{I}+({J}-{I})*{P},{A}<={J},OR({Q}="진행중",{Q}="고객검토",{Q}="수정반영"))'], stopIfTrue=True, fill=fill(C_REMAIN)),
        FormulaRule(formula=[f'AND({g},{A}>={I},{A}<={J},{Q}="대기")'], stopIfTrue=True, fill=fill(C_WAIT)),
        FormulaRule(formula=[f'AND({g},{A}>={I},{A}<={J},{Q}="보류")'], stopIfTrue=True, fill=fill(C_HOLD)),
    ]
    for rule in rules:
        ws.conditional_formatting.add(data_rng, rule)
    # 날짜 행(5)만 주말/공휴일 회색 음영
    date_rng = f"{sc}5:{lc}5"
    ws.conditional_formatting.add(date_rng, FormulaRule(
        formula=[f'AND({A}<>"",COUNTIF(공휴일!$A:$A,{A})>0)'], stopIfTrue=True, fill=fill(C_HOLIDAY)))
    ws.conditional_formatting.add(date_rng, FormulaRule(
        formula=[f'AND({A}<>"",WEEKDAY({A},2)>=6)'], stopIfTrue=True, fill=fill(C_HOLIDAY)))
    # 요일 행(6)은 검정 배경 유지, 주말/공휴일은 빨간 글씨로만 구분
    wd_rng = f"{sc}6:{lc}6"
    red = Font(name=FONT, size=8, bold=True, color="FFDC2626")
    ws.conditional_formatting.add(wd_rng, FormulaRule(
        formula=[f'AND({A}<>"",COUNTIF(공휴일!$A:$A,{A})>0)'], stopIfTrue=True, font=red))
    ws.conditional_formatting.add(wd_rng, FormulaRule(
        formula=[f'AND({A}<>"",WEEKDAY({A},2)>=6)'], stopIfTrue=True, font=red))
    # 기준일 선 (오늘=$T$3)
    ws.conditional_formatting.add(f"{sc}5:{lc}{LAST}",
                                  FormulaRule(formula=[f'{A}=$T$3'], stopIfTrue=False, fill=fill(C_TODAY)))

    # ---------- 공휴일 시트 ----------
    if "공휴일" in wb.sheetnames:
        del wb["공휴일"]
    hd = wb.create_sheet("공휴일")
    hd["A1"], hd["B1"] = "날짜", "명칭"
    for c in ("A1", "B1"):
        hd[c].font = Font(name=FONT, size=10, bold=True, color="FFFFFFFF")
        hd[c].fill = fill("FF000000")
        hd[c].alignment = Alignment(horizontal="center", vertical="center")
        hd[c].border = BORDER
    hd.column_dimensions["A"].width = 14
    hd.column_dimensions["B"].width = 24
    hd.sheet_view.showGridLines = False

    # ---------- 요약 시트 (헤더만; build_wbs가 단계별 데이터 채움) ----------
    if "요약" in wb.sheetnames:
        del wb["요약"]
    sm = wb.create_sheet("요약")
    sm.sheet_view.showGridLines = False
    sm["A1"] = "요약"
    sm["A1"].font = Font(name=FONT, size=16, bold=True)
    for c, t in {1: "단계", 2: "기간", 3: "목표진척율", 4: "진척율", 5: "상태"}.items():
        cell = sm.cell(row=3, column=c, value=t)
        cell.font = Font(name=FONT, size=10, bold=True, color="FFFFFFFF")
        cell.fill = fill("FF000000")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER
    for c, w in {1: 16, 2: 26, 3: 12, 4: 10, 5: 10}.items():
        sm.column_dimensions[get_column_letter(c)].width = w

    # ---------- 개정이력 시트 (헤더만; build_wbs가 최초 행 채움) ----------
    if "개정이력" in wb.sheetnames:
        del wb["개정이력"]
    rev = wb.create_sheet("개정이력")
    rev.sheet_view.showGridLines = False
    rev["A1"] = "개정이력"
    rev["A1"].font = Font(name=FONT, size=16, bold=True)
    for c, t in {1: "버전", 2: "날짜", 3: "작성자", 4: "변경 내용"}.items():
        cell = rev.cell(row=3, column=c, value=t)
        cell.font = Font(name=FONT, size=10, bold=True, color="FFFFFFFF")
        cell.fill = fill("FF000000")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER
    for c, w in {1: 10, 2: 14, 3: 12, 4: 55}.items():
        rev.column_dimensions[get_column_letter(c)].width = w

    # 사용안내(온보딩) 시트 추가
    build_guide_sheet(wb)

    # 시트 순서: 표지 · 사용안내 · 개정이력 · 요약 · WBS · 공휴일
    order = {"표지": 0, "사용안내": 1, "개정이력": 2, "요약": 3, "WBS": 4, "공휴일": 5}
    wb._sheets.sort(key=lambda s: order.get(s.title, 99))
    wb.active = 0
    wb.save(out)
    print("템플릿 생성(원본 수술):", out)


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("사용법: python make_template.py <원본.xlsx> <출력.xlsx>")
        sys.exit(1)
    build_template(sys.argv[1], sys.argv[2])
