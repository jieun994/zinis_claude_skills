# -*- coding: utf-8 -*-
"""
WBS 생성 스크립트 — 원본 수술 템플릿(WBS_template.xlsx)을 채운다.

컬럼(삭제 후):
  No(A~C: L1/L2/L3) | L1~L3(D~F) | 필요자료(G) | 마감기한(H) | 수행사(I) |
  담당자(J) | 산출물(K) | 시작일(L) | 종료일(M) | 진행일(N) |
  총/목표/실 작업량(O/P/Q) | 목표진척율(R) | 진척율(S) | 상태(T) | 비고(U) | 달력(V~)

규칙:
  ▸ L1은 8개 고정: 프로젝트 관리, 기획, 디자인, 퍼블리싱, 개발, 테스트, 오픈, 안정화.
  ▸ 작업량은 말단(L3)에만 입력, 상위 자동 합산: L1=ΣL2, L2=ΣL3.
  ▸ 진척율=실/총, 목표진척율=목표/총, 진행일=NETWORKDAYS — 모두 수식(자동).
  ▸ 시작/종료일로 간트 자동. 기준일=TODAY(). 공휴일 시트 자동(간트/진행일 참조).

말단 진행 입력(둘 다 가능, 기본 퍼센트): 목표진척율/진척율(0~1) 또는 목표작업량/실작업량.
사용법: python build_wbs.py <입력.json> <출력.xlsx> [템플릿.xlsx]
"""
import sys, json, datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

try:
    import holidays as _holidays
except ImportError:
    _holidays = None

L1N, L2N, L3N = 1, 2, 3
NAME = {1: 4, 2: 5, 3: 6}
필요자료, 마감기한, 수행사, 담당자, 산출물 = 7, 8, 9, 10, 11
시작일, 종료일, 진행일, 총작업량, 목표작업량, 실작업량 = 12, 13, 14, 15, 16, 17
목표진척율, 진척율, 상태, 비고 = 18, 19, 20, 21
CAL_START = 22
FIRST_ROW, LAST_ROW = 7, 66          # 60행
NDAYS = 370
L1_FIXED = ["프로젝트 관리", "기획", "디자인", "퍼블리싱", "개발", "테스트", "오픈", "안정화"]
VALID_STATUS = {"대기", "진행중", "완료", "보류", "고객검토", "수정반영"}
WD = ["월", "화", "수", "목", "금", "토", "일"]
FONT = "Malgun Gothic"
thin = Side(style="thin", color="FFD1D5DB")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def fill(c):
    return PatternFill("solid", fgColor=c)


def to_date(s):
    if not s:
        return None
    if isinstance(s, datetime.datetime):
        return s
    if isinstance(s, datetime.date):
        return datetime.datetime(s.year, s.month, s.day)
    return datetime.datetime.strptime(str(s)[:10], "%Y-%m-%d")


def r2(x):
    return None if x is None else round(float(x), 2)


def kr_holidays(start, end):
    if _holidays is None:
        print("  [경고] holidays 라이브러리 없음 → 공휴일 비움 (pip install holidays)")
        return []
    kr = _holidays.SouthKorea(years=list(range(start.year, end.year + 1)))
    return [(datetime.datetime(d.year, d.month, d.day), kr[d]) for d in sorted(kr)
            if start.date() <= d <= end.date() and "제헌절" not in kr[d]]


def networkdays(start, end, hol):
    """평일(월~금) 중 공휴일 제외 일수. (엑셀 NETWORKDAYS와 동일)"""
    if not start or not end or end < start:
        return 0
    n, d = 0, start
    while d <= end:
        if d.weekday() < 5 and d.date() not in hol:
            n += 1
        d += datetime.timedelta(days=1)
    return n


def compute_structure(tasks, hol):
    """직속 자식·부모 일정(min/max)·상태 추정값 계산. (실제 셀은 수식으로 들어감)"""
    n = len(tasks)
    levels = [int(t.get("level", 1)) for t in tasks]
    children = [[] for _ in range(n)]
    stack = []
    for i in range(n):
        while stack and stack[-1][0] >= levels[i]:
            stack.pop()
        if stack:
            children[stack[-1][1]].append(i)
        stack.append((levels[i], i))

    for i in sorted(range(n), key=lambda i: levels[i], reverse=True):
        t = tasks[i]
        t["_children"] = children[i]
        if not children[i]:                       # 말단(L3)
            t["_start"], t["_end"] = to_date(t.get("시작일")), to_date(t.get("종료일"))
            R = networkdays(t["_start"], t["_end"], hol)   # 총작업량=진행일
            prog = float(t.get("진척율") or 0)
            t["_R"], t["_T"] = R, R * prog                 # 상태 추정용
        else:                                     # 상위(L1/L2)
            kids = [tasks[c] for c in children[i]]
            t["_R"] = sum(k["_R"] for k in kids)
            t["_T"] = sum(k["_T"] for k in kids)
            starts = [k["_start"] for k in kids if k.get("_start")]
            ends = [k["_end"] for k in kids if k.get("_end")]
            t["_start"] = to_date(t.get("시작일")) or (min(starts) if starts else None)
            t["_end"] = to_date(t.get("종료일")) or (max(ends) if ends else None)
        st = t.get("상태")
        if not st:
            v = (t["_T"] / t["_R"]) if t["_R"] else 0
            st = "완료" if v >= 1 else ("진행중" if v > 0 else "대기")
        t["_status"] = st
    return tasks


def write_summary(wb, l1_rows):
    """요약 시트: 전체(통합) + 단계(L1)별 — WBS 셀과 수식 연결(자동 갱신)."""
    sm = wb["요약"]
    Osum = "+".join(f"WBS!O{r}" for r in l1_rows) or "0"   # 총작업량
    Psum = "+".join(f"WBS!P{r}" for r in l1_rows) or "0"   # 목표작업량
    Qsum = "+".join(f"WBS!Q{r}" for r in l1_rows) or "0"   # 실작업량
    # 전체(통합) 행 (4행)
    sm["A4"] = "전체"
    sm["B4"] = '=TEXT(WBS!$M$3,"yyyy-mm-dd")&" ~ "&TEXT(WBS!$R$3,"yyyy-mm-dd")'
    sm["C4"] = f'=IFERROR(({Psum})/({Osum}),"")'
    sm["D4"] = f'=IFERROR(({Qsum})/({Osum}),"")'
    sm["E4"] = '=IF(D4="","",IF(D4>=1,"완료",IF(D4>0,"진행중","대기")))'
    # 단계(L1)별 행 (5행~) — WBS 값을 그대로 참조
    for i, wr in enumerate(l1_rows):
        r = 5 + i
        sm.cell(row=r, column=1).value = f"=WBS!D{wr}"
        sm.cell(row=r, column=2).value = (
            f'=TEXT(WBS!L{wr},"yyyy-mm-dd")&" ~ "&TEXT(WBS!M{wr},"yyyy-mm-dd")')
        sm.cell(row=r, column=3).value = f"=WBS!R{wr}"
        sm.cell(row=r, column=4).value = f"=WBS!S{wr}"
        sm.cell(row=r, column=5).value = f"=WBS!T{wr}"
    # 서식
    last = 4 + len(l1_rows)
    for r in range(4, last + 1):
        for c in range(1, 6):
            cell = sm.cell(row=r, column=c)
            cell.border = BORDER
            cell.font = Font(name=FONT, size=10, bold=(r == 4))
            cell.alignment = Alignment(horizontal="left" if c == 1 else "center",
                                       vertical="center")
            if r == 4:
                cell.fill = fill("FFF3F4F6")
        sm.cell(row=r, column=3).number_format = "0.0%"
        sm.cell(row=r, column=4).number_format = "0.0%"


def build(input_json, output_xlsx, template_xlsx):
    with open(input_json, encoding="utf-8") as f:
        data = json.load(f)
    proj = data.get("project", {})
    tasks = data.get("tasks", [])
    if len(tasks) > (LAST_ROW - FIRST_ROW + 1):
        raise SystemExit(f"작업이 너무 많습니다(최대 {LAST_ROW-FIRST_ROW+1}개).")
    # 시작에 받는 값: 프로젝트명 / 시작일 / 종료일 (기준일은 오늘 자동)
    start = to_date(proj.get("start"))
    end = to_date(proj.get("end"))
    if not start or not end:
        raise SystemExit("project.start / project.end (시작일·종료일) 필수.")

    # 공휴일 먼저 계산 (총작업량=진행일 계산에 필요)
    hol = data.get("holidays")
    if hol:
        pairs = [(to_date(h[0]), h[1] if len(h) > 1 else "") if isinstance(h, (list, tuple))
                 else (to_date(h), "") for h in hol]
    else:
        pairs = kr_holidays(start, end)
    hol_set = set(d.date() for d, _ in pairs)

    compute_structure(tasks, hol_set)

    wb = openpyxl.load_workbook(template_xlsx)
    ws = wb["WBS"]

    # 메타: 시작 M3, 종료 R3, 기준일 T3=오늘(TODAY) 자동
    ws["M3"] = start; ws["M3"].number_format = "yyyy-mm-dd"
    ws["R3"] = end;   ws["R3"].number_format = "yyyy-mm-dd"
    ws["T3"] = "=TODAY()"; ws["T3"].number_format = "yyyy-mm-dd"

    # 표지
    cv = wb["표지"]
    cv["B11"] = proj.get("name", "")
    if proj.get("start") and proj.get("end"):
        cv["D18"] = f"{proj['start']} ~ {proj['end']}"
    cv["D20"] = proj.get("author", "")
    if proj.get("author_date"):
        cv["D22"] = to_date(proj["author_date"]); cv["D22"].number_format = "yyyy-mm-dd"
    cv["D24"] = proj.get("version", "")

    # 개정이력 시트 — 최초 작성 행
    rev = wb["개정이력"]
    rev["A4"] = proj.get("version", "")
    if proj.get("author_date"):
        rev["B4"] = to_date(proj["author_date"]); rev["B4"].number_format = "yyyy-mm-dd"
    rev["C4"] = proj.get("author", "")
    rev["D4"] = "최초 작성"
    for c in range(1, 5):
        cell = rev.cell(row=4, column=c)
        cell.border = BORDER
        cell.font = Font(name=FONT, size=10)
        cell.alignment = Alignment(horizontal="left" if c == 4 else "center", vertical="center")

    # 공휴일 시트 (간트/진행일이 공휴일!$A:$A 참조)
    hd = wb["공휴일"]
    for r in range(2, 120):
        hd.cell(row=r, column=1).value = None
        hd.cell(row=r, column=2).value = None
    for i, (d, name) in enumerate(pairs):
        hd.cell(row=2 + i, column=1).value = d
        hd.cell(row=2 + i, column=1).number_format = "yyyy-mm-dd"
        hd.cell(row=2 + i, column=2).value = name

    # 선택 칸(필요자료/마감기한/수행사) 표시 여부 — 기본 숨김, 요청한 것만 표시
    OPTIONAL = {"필요자료": 7, "마감기한": 8, "수행사": 9}
    wanted = set(data.get("optional_columns") or [])
    for nm, col in OPTIONAL.items():
        ws.column_dimensions[get_column_letter(col)].hidden = (nm not in wanted)

    # 달력 날짜는 템플릿 수식($M$3/$R$3 참조)이 자동 생성.
    days = (end - start).days + 1
    if days > NDAYS:
        raise SystemExit(f"기간이 너무 깁니다({days}일 > 최대 {NDAYS}일).")
    # 종료일 이후 열만 숨겨 간트 길이를 프로젝트 기간에 맞춘다.
    for col in range(CAL_START, CAL_START + NDAYS):
        ws.column_dimensions[get_column_letter(col)].hidden = (col >= CAL_START + days)
    # 간트 격자선(보이는 달력 칸 전체) — 흐린 hair 대신 thin 실선
    gside = Side(style="thin", color="FFD9D9D9")
    gborder = Border(left=gside, right=gside, top=gside, bottom=gside)
    for col in range(CAL_START, CAL_START + days):
        for r in range(5, LAST_ROW + 1):
            ws.cell(row=r, column=col).border = gborder

    # 행4 월 라벨: 월별로 셀 병합 (라벨이 잘리지 않게)
    cur = (start.year, start.month); m_start = CAL_START
    for i in range(days):
        dd = start + datetime.timedelta(days=i)
        if (dd.year, dd.month) != cur:
            if CAL_START + i - 1 >= m_start:
                ws.merge_cells(start_row=4, start_column=m_start,
                               end_row=4, end_column=CAL_START + i - 1)
            cur = (dd.year, dd.month); m_start = CAL_START + i
    ws.merge_cells(start_row=4, start_column=m_start,
                   end_row=4, end_column=CAL_START + days - 1)

    # L1 고정 검증
    for t in tasks:
        if int(t.get("level", 1)) == 1 and t.get("name") not in L1_FIXED:
            print(f"  [주의] L1 '{t.get('name')}' 은 고정 8종({L1_FIXED})에 없음.")

    a = c = d = 0
    l1_rows = []
    for i, t in enumerate(tasks):
        r = FIRST_ROW + i
        lvl = int(t.get("level", 1))
        if lvl == 1:
            a += 1; c = d = 0; l1_rows.append(r)
        elif lvl == 2:
            c += 1; d = 0
        else:
            d += 1
        # No: 각 자릿수가 레벨 (L1=a00, L2=ac0, L3=acd). 빈 자리는 0.
        ws.cell(row=r, column=L1N).value = a
        ws.cell(row=r, column=L2N).value = c
        ws.cell(row=r, column=L3N).value = d
        ws.cell(row=r, column=NAME[lvl]).value = t.get("name", "")

        # 레벨별 행 색 (L1 짙은 회색/L2 연노랑/L3 흰색)
        row_fill = {1: "FF595959", 2: "FFFFF2CC", 3: "FFFFFFFF"}[lvl]
        row_font = Font(name=FONT, size=9, bold=(lvl == 1),
                        color="FFFFFFFF" if lvl == 1 else "FF000000")
        for cc in range(1, 22):           # 표 영역(A~U)만, 달력은 제외
            cell = ws.cell(row=r, column=cc)
            cell.fill = fill(row_fill)
            cell.font = row_font

        def setv(col, val):
            if val is not None and val != "":
                ws.cell(row=r, column=col).value = val
        setv(필요자료, t.get("필요자료"))
        if t.get("마감기한"):
            cc = ws.cell(row=r, column=마감기한); cc.value = to_date(t["마감기한"]); cc.number_format = "yyyy-mm-dd"
        setv(수행사, t.get("수행사"))
        setv(담당자, t.get("담당자"))
        setv(산출물, t.get("산출물"))
        if t.get("_start"):
            cc = ws.cell(row=r, column=시작일); cc.value = t["_start"]; cc.number_format = "yyyy-mm-dd"
        if t.get("_end"):
            cc = ws.cell(row=r, column=종료일); cc.value = t["_end"]; cc.number_format = "yyyy-mm-dd"
        st = t["_status"]
        if st not in VALID_STATUS:
            print(f"  [경고] {r}행 상태 '{st}' 표준값 아님.")
        ws.cell(row=r, column=상태).value = st
        setv(비고, t.get("비고"))

        # ----- 자동 계산 수식 -----
        childrows = [FIRST_ROW + ci for ci in t["_children"]]
        is_leaf = not childrows
        HD = "공휴일!$A$2:$A$80"
        # 진행일 N = NETWORKDAYS(시작, 종료, 공휴일)
        ws.cell(row=r, column=진행일).value = \
            f'=IF(OR(L{r}="",M{r}=""),"",NETWORKDAYS(L{r},M{r},{HD}))'
        # 총작업량 O : L3=진행일, 상위=직속 자식 합
        ws.cell(row=r, column=총작업량).value = \
            f'=N{r}' if is_leaf else "=" + "+".join(f"O{cr}" for cr in childrows)
        # 목표작업량 P : 총 * (기준일 vs 시작/종료 가중치)
        ws.cell(row=r, column=목표작업량).value = (
            f'=O{r}*IF($T$3<L{r},0,IF($T$3>=M{r},1,'
            f'IFERROR(NETWORKDAYS(L{r},$T$3,{HD})/NETWORKDAYS(L{r},M{r},{HD}),0)))')
        # 실작업량 Q : L3=총*진척율, 상위=직속 자식 합
        ws.cell(row=r, column=실작업량).value = \
            f'=O{r}*S{r}' if is_leaf else "=" + "+".join(f"Q{cr}" for cr in childrows)
        # 목표진척율 R = 목표/총, 진척율 S : L3=입력값, 상위=실/총
        ws.cell(row=r, column=목표진척율).value = f'=IFERROR(P{r}/O{r},"")'
        if is_leaf:
            ws.cell(row=r, column=진척율).value = round(float(t.get("진척율") or 0), 4)
        else:
            ws.cell(row=r, column=진척율).value = f'=IFERROR(Q{r}/O{r},"")'
        # 표시 형식
        for col in (진행일, 총작업량, 목표작업량, 실작업량):
            ws.cell(row=r, column=col).number_format = "0.0"
        for col in (목표진척율, 진척율):
            ws.cell(row=r, column=col).number_format = "0.0%"

    # (#4) 작업 이후 빈 행은 흰색으로 (표 격자 테두리는 60행까지 유지)
    for r in range(FIRST_ROW + len(tasks), LAST_ROW + 1):
        for cc in range(1, 22):
            cell = ws.cell(row=r, column=cc)
            cell.value = None
            cell.fill = fill("FFFFFFFF")
            cell.border = BORDER
            cell.font = Font(name=FONT, size=9)

    # 요약 시트 채우기 (전체 + 단계별)
    write_summary(wb, l1_rows)

    wb.save(output_xlsx)
    print(f"완료: {output_xlsx}  (작업 {len(tasks)}개, 공휴일 {len(pairs)}개)")


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("사용법: python build_wbs.py <입력.json> <출력.xlsx> [템플릿.xlsx]")
        sys.exit(1)
    tpl = sys.argv[3] if len(sys.argv) > 3 else "template/WBS_template.xlsx"
    build(sys.argv[1], sys.argv[2], tpl)
